# -*- coding: utf-8 -*-            
# @Author : libeijie
# @Time : 2023/9/12 13:59

import os
import io
import logging
import xlrd
import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation

from Salvation.settings import BASE_DIR
from common.cfg_helper import *
from common.excel_helper import LoadExcelData

logger = logging.getLogger('salvation')

class ReuseCommonTestCases:
    def __init__(self):
        #self.formwork_path = '配置文件'
        self.testcase_filename = '' #读取目标excel名
        self.extract_sheetname = '用例设计' #读取目标excel的sheet页的名
        self.extract_rowbegin = 1 #读取的起始行
        self.extract_sheetname_gn = {'前端功能页面':'前端功能页面配置'}
        self.extract_sheetname_jkfw = {'对外接口服务':'对外接口服务配置'}
        self.extract_rowbegin_gn = 1 #读取“前端功能页面配置”页的起始行，从0开始
        self.extract_colbegin_gn = 15 #读取“前端功能页面配置”页的起始列，从0开始
        self.extract_colfinish_gn = 34 #读取“前端功能页面配置”页的结束列，从0开始 （我数的是33？？？？）
        self.extract_rowbegin_jkfw = 1 #读取“对外接口服务”页的起始行，从0开始
        self.extract_colend_jkfw = 10 #读取“对外接口服务”页的结束列，从0开始 （我数的是3？？？？）
        self.extract_name_gn = 4 #“前端功能页面配置”页 的 功能名称+数据等级 的列号
        self.delete_sheetname_list = ['执行单元-示例1','执行单元-示例2']
        cfg_file = ''.join([str(BASE_DIR), '/apps/data/config.cfg'])
        self.cfg_read = CfgHelper(cfg_file)
        self.formwork_path = self.cfg_read.read_cfg_get_value_by_key('config_file', 'file_config_path')

    def create_project_test_cases(self,file):
        '''
        生成用例
        :param file: 上传的用例模板文件
        :return:
        '''
        # 获取excel对象
        # 将文件内容存储在内存中，防止多次读取上传文件指针移至末尾报错
        file_content = io.BytesIO(file.read())
        # 读取测试用例文档
        extract_excel = LoadExcelData()


        workbook = load_workbook(file,keep_vba=True) #如果file中没有vba，则使用此方法会导致file被损坏

        # 获取“用例设计”sheet页的二维数据列表，从第2行开始
        data_list = extract_excel.load_from_excel_by_upload(file_content,self.extract_sheetname,self.extract_rowbegin)
        # 读取“前端功能页面配置”sheet页的二维数据列表，从第2行开始
        data_gn_list = extract_excel.load_from_excel_by_upload(file_content, self.extract_sheetname_gn['前端功能页面'],self.extract_rowbegin_gn)
        # 获取“对外接口服务配置”sheet页的二维数据列表，从第2行开始
        data_jkfw_list = extract_excel.load_from_excel_by_upload(file_content, self.extract_sheetname_jkfw['对外接口服务'],self.extract_rowbegin_jkfw)
        if not data_list:
            message = '文件异常，请检查文档sheet页[%s]是否为最新模板！' % (self.extract_sheetname)
            logger.error(message)
            return {'code':1,'message':message}


        xl = ExcelExtractMergeSheet()
        xl.read_excel_tag(workbook)

        for row in data_list:
            module = row[0]  # “用例设计”sheet页的”模块编码”列
            test_type = row[1]  # “用例设计”sheet页的“测试点分类”列
            test_flg = row[2]  # “用例设计"sheet页的"项目是否涉及”列
            if (module =='None' or module =='') and (test_type =='None' or test_type =='') and (test_flg =='None' or test_flg ==''):
                break
            if test_flg != 'None' and test_flg != '':
                logger.info('要合并测试点: %s' % test_type)
                # 拷贝sheet页内容
                # 当”测试点分类”为前端功能页面时:
                # 读取“前端功能页面配置”sheet页相关功能的控件，与配置文件中的《前端功能页面.xlsm》中相关sheet页匹配
                # 将匹配到的用例通过调用函数extract_multisheet写至指定的sheet页
                gn_count = 0
                if test_type in self.extract_sheetname_gn:  # 前端功能页面
                    if not data_gn_list:
                        message = '文件异常，请检查文档sheet页[%s]是否为最新模板!' %(self.extract_sheetname_gn['前端功能页面'])
                        logger.error(message)
                        return {'code':1,'message':message}
                    gn_name_list = data_gn_list[0][self.extract_colbegin_gn: self.extract_colfinish_gn + 1] #['查询时点', '查询时段', '单项输入框', '多项输入框', '左右互选框', '多项选择框', '单选下拉列表', '多选下来列表', '复选框', '单选按钮', 'tabs标签', '自定义表单', '自定义选择框', '查询按钮', '重置按钮', '上传按钮', '下载', '其他按钮', '查询报表']
                    for data_row in range(1, len(data_gn_list)):  # 获取最新二维数据列表的行数
                        if not data_gn_list[data_row][self.extract_name_gn]: # 功能名称+数据等级 如果为空，则循环终止
                            break
                        testcase_sheetname = data_gn_list[data_row][self.extract_name_gn] #功能名称+数据等级，例如：结算信息汇总查询-沪（Ⅰ级）、资金账余额变动查询_深（Ⅰ级）等
                        logger.info('功能名称+数据等级:%s' % testcase_sheetname)
                        gn_flg_list = data_gn_list[data_row][self.extract_colbegin_gn:self.extract_colfinish_gn + 1] #['清算日期', '挂牌年份', 'None', '交收会员\n结果代码\n结算单元\n证券账户\n证券代码Ⅰ\n资金账号', '记录类型\n交收方式\n交易方式\n清算标志\n市场代码\n业务类型', '流通类型\n买卖标志\n权益类别\n账户性质\n证券类别', 'None', 'None', 'None', 'None', 'None', 'None', '查询', '重置', 'None', 'None', 'None', 'Cognos报表']
                        tmp_list = [['前端功能页面-基本验证', {'D': testcase_sheetname}]]
                        for num in range(len(gn_flg_list)):
                            if gn_flg_list[num] and gn_flg_list[num] != 'None' and gn_flg_list[num] != '':
                                logger.info('待复制的控件: %s %s' % (gn_name_list[num], gn_flg_list[num].split('\n')))
                                #当“前端功能页面配置”sheet页中的表头为“查询报表”时，通过“查询报表”列值name在配置文件中匹配对应的sheet页
                                if gn_name_list[num] == '查询报表':
                                    for name in gn_flg_list[num].split('\n'):
                                        tmp_list.append([name, {'D': testcase_sheetname, 'E': name}]) #[['Cognos报表', {'D': '结算信息汇总查询-沪（Ⅰ级）', 'E': 'Cognos报表'}]]
                                else:
                                    # 当页面中的控件不为“查询报表”时，通过gn_name_list[num]在配置文件中匹配对应的sheet页
                                    for name in gn_flg_list[num].split('\n'):
                                        tmp_list.append([gn_name_list[num],{'D':testcase_sheetname, 'E':'{0}-{1}'.format(gn_name_list[num], name)}]) #[['左右互选框', {'D': '结算信息汇总查询-沪（Ⅰ级）', 'E': '左右互选框-交收会员'}], ['左右互选框', {'D': '结算信息汇总查询-沪（Ⅰ级）', 'E': '左右互选框-结果代码'}], ['左右互选框', {'D': '结算信息汇总查询-沪（Ⅰ级）', 'E': '左右互选框-结算单元'}]]

                        xl.extract_multisheet(self.formwork_path,test_type, testcase_sheetname,tmp_list)
                        # ======================================修改功能sheet页的用例编号===========================================
                        # 定位内存文件中的目标文件的sheet页
                        ws_tag = xl.tag_excelbook[testcase_sheetname] #用例模板文件 结算信息汇总查询-沪（Ⅰ级）、资金账余额变动查询_深（Ⅰ级）等 sheet页
                        gn_count += 1
                        self.modify_testcases_id(ws_tag, 'QDGN', gn_count) #对第一列赋值为QDGN1-0001、QDGN1-0002...QDGN4-0001、QDGN4-0002
                        #调整整列值
                        self.modify_testcase_mergename(ws_tag, 11, '''=IF(K%s="自动化",E%s,"") ''', [0, 0],{'4':'全流程;界面基本功能'})
                        # self.modify_testcase_mergename(ws_tag, 11,'''=IF(K%s="自动化",IF(E%s="全流程","流程类",IF(E%s="界面基本功能","静态类",E%s)),"") ''',
                        #                                [0, 0,0,0],{'4':'全流程；界面基本功能'}) #L列（TestSuite列赋值，如果K是自动化，则L列取E列值）

                        # =====================================================================================================
                        if not xl.conditional_formatting(testcase_sheetname):  # 设置条件格式

                            return False
                # 当“测试点分类”为对外接口服务时
                elif test_type in self.extract_sheetname_jkfw:
                    # 读取“对外接口服务配置”页的接口对象类别(宏、视图、表、存储过程、函数等)
                    if not data_jkfw_list:
                        logger.error('文件异常，请检查文档sheet页[%s]是否为最新模板!'%(self.extract_sheetname_jkfw['对外接口服务']))
                        return False
                    jkfw_name_list = data_jkfw_list[0][:self.extract_colend_jkfw + 1]
                    macro_count = 0
                    # 获取当前sheet页的总行数
                    for data_col in range(0, len(jkfw_name_list)):
                        for data_row in range(1, len(data_jkfw_list)):
                            sheet_name = "对外接口-" + data_jkfw_list[0][data_col]  # 要创建和读取的sheet页名称
                            if data_jkfw_list[0][data_col] == '宏':
                                if data_jkfw_list[data_row][data_col]:
                                    jkfw_macro_dict = {'D': data_jkfw_list[data_row][data_col]}
                                    sheet_name_new = data_jkfw_list[data_row][data_col].split('.')[1]
                                    xl.extract_sheet(self.formwork_path, test_type,sheet_name_new, sheet_name, replace_dict=jkfw_macro_dict)
                                    # =============================修改宏编号 (20200208)==================================
                                    # 定位内存文件中的目标文件的sheet页
                                    ws_tag = xl.tag_excelbook[sheet_name_new]
                                    # #根据序号修改宏用例编号 (JKFW M-001 ->JKFW M1-001)
                                    macro_count += 1
                                    self.modify_testcases_id(ws_tag, 'JKFW_M', macro_count)
                                    # ===================================================================================
                                    logger.info(data_jkfw_list[0][data_col] + data_jkfw_list[data_row][data_col] + '测试用例生成完成!')
                                    if not xl.conditional_formatting(data_jkfw_list[data_row][data_col].split('.')[1]): #设置条件格式
                                        return False
                                elif data_row == 1:
                                    logger.info('接口服务中' + data_jkfw_list[0][data_col] + '为空，请确认项目是否涉及，若涉及则在相关配置页补充；若不涉及留空即可')
                                    break
                                else:
                                    break
                            else:
                                if data_jkfw_list[data_row][data_col]:
                                    jkfw_vt_dict = {'D': data_jkfw_list[data_row][data_col]}
                                    xl.extract_sheet(self.formwork_path, test_type, sheet_name,sheet_name, remove_flg=False, replace_dict=jkfw_vt_dict)
                                    logger.info(data_jkfw_list[0][data_col] + data_jkfw_list[data_row][data_col] + '测试用例生成完成!')
                                    if not xl.conditional_formatting(sheet_name): # 设置条件格式
                                        return False
                                elif data_row == 1:
                                    logger.info('接口服务中'+ data_jkfw_list[0][data_col] + '为空, 请确认项目是否涉及。若涉及则在相关配置页补充: 若不涉及空即可')
                                else:
                                    break
                else:
                    xl.extract_sheet(self.formwork_path, test_type, test_type, test_type)
                    if not xl.conditional_formatting(test_type):  # 设置条件格式
                        return False
        for delete_sheetname in self.delete_sheetname_list:
            xl.delete_sheet(delete_sheetname)
        workbook.close()
        return workbook

    def modify_testcases_id(self, worksheet, case_type, num):
        '''
        对sheet页第一列赋值
        :param worksheet: 读取到内存文件中的sheet页(含表头)
        :param case_type: 要修改的用例分类(口(JKFW_M)、前端功能（QDGN）等）
        :param num: 分类中相应的接口、功能的序号
        :return:
        '''
        for i in range(1, worksheet.max_row): #循环行
            index_len = 3  # 用例编号的数字的长度
            # 序号不为3位数的则拿0补全
            if len(str(i)) <= index_len:
                caseid = "{}".format('0' * (3 - len(str(i)))) + str(i)
                new_value = case_type + str(num) + '-' + caseid
                # 使用修改后的值替换原有值
                worksheet[i + 1][0].value = new_value

    def modify_testcase_mergename(self, worksheet, modify_col_num, modify_cont, excur_list, exclude=None, modify_start_row=1):
        '''
        调整测试用例中某整列的内容
        :param worksheet: 读取到内存文件中的sheet页(含表头)
        :param modify_col_num: 修改列号
        :param modify_cont: 修改内容
        :param excur_list: 修改内容中的偏移量列表
        :param modify_start_row: 修改的起始行
        :return:
        '''

        for i in range(modify_start_row + 1, worksheet.max_row+1):
            if exclude: #如果有排除项exclude {'4':'全流程；界面基本功能'}
                exclude_len = len(exclude)
                exclude_true_num = 0
                #如果此列值在exclude中，则不做处理
                for exclude_col in exclude.keys(): #4 即E列
                    exclude_col_value = exclude[exclude_col].split(';')  # ['全流程','界面基本功能']
                    if worksheet[i][int(exclude_col)].value in exclude_col_value: #如果Ei值是'全流程；界面基本功能'，则这一行都不做任何更改
                        exclude_true_num +=1
                        continue
                if exclude_true_num == exclude_len: #此行exclude的key列 都被满足
                    continue
                else: # 不满足排除条件，那就按原流程来
                    self.modify_testcase_cell_value(excur_list,i,modify_cont,worksheet,modify_col_num)
            else:
                self.modify_testcase_cell_value(excur_list,i,modify_cont,worksheet,modify_col_num)

    def modify_testcase_cell_value(self,excur_list,i,modify_cont,worksheet,modify_col_num):
        '''
        处理某个单元格的值
        :param excur_list: 修改内容中的偏移量列表
        :param i: 行
        :param modify_cont: 修改内容
        :param worksheet: 读取到内存文件中的sheet页(含表头)
        :param modify_col_num: 列
        :return:
        '''
        tmp_excur_list = []
        for x in excur_list:
            tmp_excur_list.append(int(x) + i)
        excur_tuple = tuple(tmp_excur_list)
        new_value = modify_cont % (excur_tuple)
        worksheet[i][modify_col_num].value = new_value


    def get_cells_by_sheetname_for_read(self,workbook,sname):
        '''
        处理file，获取excel中sname sheet页的数据
        :param workbook:传入的excel对象
        :param sname:sheet名称
        :return: 返回4元素元组（取值矩阵，类型矩阵，总行数，总列数）
        '''
        sheet_names_list = workbook.sheetnames
        if sname not in sheet_names_list:
            return None, None, None, None
        try:
            excelsheet = workbook.get_sheet_by_name(sname)
        except Exception as e:
            logger.error('报错信息：' + repr(e) + ',报错的行号:' + str(e.__traceback__.tb_lineno))
            return None, None, None, None

        nr = excelsheet.max_row #有效行数
        nc = excelsheet.max_column #有效果列数
        if nr == 0 or nc == 0:
            logger.warning('读取到的有效行数是[%s](最大5000行)，有效列数是[%s](最大100列)。请检查excel文档：', nr, nc)
            return None, None, None, None
        if nr >= 5000 or nc >= 100:
            logger.warning('读取到的有效行数是[%s](最大5000行)，有效列数是[%s](最大100列)已经超出正常范围！请检查excel文档：',nr,nc)
            return None, None, None, None

        valuematrix = []
        ctypematrix = []
        for x in range(1,nr+1): #循环行
            valuematrix.append([])
            ctypematrix.append([])
            for y in range(1,nc+1): #循环列
                #无法获得确切行数列数，取数会出错，可忽略异常，把所有能取出的数抽取
                try:
                    # 根据行列值获取单元格对象
                    cell = excelsheet.cell(row=x, column=y)
                    # 获取单元格的数据类型
                    cty = cell.data_type #s
                    val = cell.value

                    if cty == 'n' and (str(val) != 'None') and str(val).split('.')[1] == '0': #number类型 小数位如果是‘.0’ 就需要去掉
                        val = str(val).split('.'[0])
                    else:
                        val = str(val).strip()
                except:
                    val = ''
                    cty = 6
                    pass
                valuematrix[x - 1].append(val)
                ctypematrix[x - 1].append(cty)
        return valuematrix,ctypematrix,nr,nc


    def load_from_excel(self,workbook,sname,rowbegin):
        '''
        获取excel中sheet页的数据
        :param file:上传的文件
        :param sname:sheet页名
        :param rowbegin:开始读取的开始行号（从0开始）
        :return:数据二维列表（为None表示读取失败）
        '''

        self.data_list = None
        (vm, cm, nr, nc) = self.get_cells_by_sheetname_for_read(workbook,sname)
        if not vm:
            return None
        self.data_list = copy.deepcopy(vm[rowbegin:][:])
        return self.data_list

class  ExcelExtractMergeSheet(object):
    '''
    从源文档拷贝指定sheet页所有内容（包括sheet页名称）到目标文档
    '''
    def __init__(self):
        self.src_excelbook = None #用例配置文件，这个在服务器上
        self.tag_excelbook = None #用例模板文件，这个是上传的
        self.rowbegin_tag = 0

        dxf_bg_brown =DifferentialStyle(fill=PatternFill(start_color='FF0000',end_color='FFC000')) #单元格背景色为‘FFC000’（棕色）
        dxf_bg_green = DifferentialStyle(fill=PatternFill(start_color='92D050',end_color='92D050'))#单元格背景色为‘92D050’（绿色）
        dxf_bg_yellow = DifferentialStyle(fill=PatternFill(start_color='FFFF00',end_color='FFFF00'))#单元格背景色为‘FFFF00’（黄色）
        dxf_bg_red = DifferentialStyle(fill=PatternFill(start_color='FF0000',end_color='FF0000'))#单元格背景色为‘FF0000’（红色）
        dxf_font_bold_red = DifferentialStyle(font=Font(bold=True,color='FF0000')) #字体为粗体，红色
        # dxf_font_black = DifferentialStyle(font=Font(color='000000')) #字体黑色
        dxf_font_black_not_italic = Font(italic = False,color='000000') #字体黑色，去除斜体

        #设置条件格式：单元格值=‘手工’
        # self.rule_cell1 = Rule(type='cellIs',operator='equal',dxf=dxf_bg_brown,formula=['手工'])
        # #设置条件格式：单元格值=‘自动化’
        # self.rule_cell2 = Rule(type='cellIs',operator='equal',dxf=dxf_bg_green,formula=['自动化'])
        # # 设置条件格式：单元格值=‘通过’
        # self.rule_cell3 = Rule(type='cellIs',operator='equal',dxf=dxf_bg_green,formula=['通过'])
        # # 设置条件格式：单元格值=‘不适用’
        # self.rule_cell4 = Rule(type='cellIs',operator='equal',dxf=dxf_bg_yellow,formula=['不适用'])
        # # 设置条件格式：单元格值=‘失败’
        # self.rule_cell5 = Rule(type='cellIs',operator='equal',dxf=dxf_bg_red,formula=['失败'])
        # # 设置条件格式：单元格值=‘低’
        # self.rule_cell6 = Rule(type='cellIs',operator='equal',dxf=dxf_bg_green,formula=['低'])
        # # 设置条件格式：单元格值=‘中’
        # self.rule_cell7 = Rule(type='cellIs',operator='equal',dxf=dxf_bg_brown,formula=['中'])
        # # 设置条件格式：单元格值=‘高’
        # self.rule_cell8 = Rule(type='cellIs',operator='equal',dxf=dxf_bg_red,formula=['高'])
        # # 设置条件格式：单元格值=‘反’
        # self.rule_cell9 = Rule(type='cellIs',operator='equal',dxf=dxf_font_bold_red,formula=['反'])
        # #设置条件格式
        # self.rule_cell10 = Rule(type='notContainsBlanks',dxf=dxf_font_black)


        self.rule_cell1 = Rule(type='containsText', operator='containsText', dxf=dxf_bg_brown,formula=['NOT(ISERROR(SEARCH("手工",{0})))'])
        self.rule_cell2 = Rule(type='containsText', operator='containsText', dxf=dxf_bg_green,formula=['NOT(ISERROR(SEARCH("自动化",{0})))'])
        self.rule_cell3 = Rule(type='containsText', operator='containsText', dxf=dxf_bg_green,formula=['NOT(ISERROR(SEARCH("通过",{0})))'])
        self.rule_cell4 = Rule(type='containsText', operator='containsText', dxf=dxf_bg_yellow,formula=['NOT(ISERROR(SEARCH("不适用",{0})))'])
        self.rule_cell5 = Rule(type='containsText', operator='containsText', dxf=dxf_bg_red,formula=['NOT(ISERROR(SEARCH("失败",{0})))'])
        self.rule_cell6 = Rule(type='containsText', operator='containsText', dxf=dxf_bg_green,formula=['NOT(ISERROR(SEARCH("低",{0})))'])
        self.rule_cell7 = Rule(type='containsText', operator='containsText', dxf=dxf_bg_brown,formula=['NOT(ISERROR(SEARCH("中",{0})))'])
        self.rule_cell8 = Rule(type='containsText', operator='containsText', dxf=dxf_bg_red,formula=['NOT(ISERROR(SEARCH("高",{0})))'])
        self.rule_cell9 = Rule(type='containsText', operator='containsText', dxf=dxf_font_bold_red,formula=['NOT(ISERROR(SEARCH("反",{0})))'])
        # self.rule_cell10 = Rule(type='containsText', operator='containsText', dxf=dxf_font_black,formula=['ISERROR(SEARCH("$",K1))'])
        self.font_cell1 = dxf_font_black_not_italic


    def read_excel_src(self,src_file):
        '''

        :param src_file: 带路径的源excel文件（没有后缀）
        :return: True/False
        '''
        if os.path.exists(src_file+'.xlsx'):
            src_file = src_file+'.xlsx'
        elif os.path.exists(src_file+'.xlsm'):
            src_file = src_file + '.xlsm'
        else:
            logger.warning('源文件[%s.xlsx]或[%s.xlsm]不存在：',src_file)
            return False
        try:
            self.src_excelbook = load_workbook(src_file)
        except Exception as e:
            logger.error('报错信息：' + repr(e) + ',报错的行号:' + str(e.__traceback__.tb_lineno))
            return False
        return True

    def read_excel_tag(self,tag_file):
        '''

        :param tag_file: excel文件
        :return: True/False
        '''
        self.tag_excelbook = tag_file



    def extract_multisheet(self,formwork_path,formwork_name,sheet_name_tag,sheet_name_list):
        '''
        多个sheet页合并成一个sheet页
        :param formwork_path: 配置文件文档路径
        :param formwork_name: 配置文件名称
        :param sheet_name_tag: 要创建的sheet页名称
        :param sheet_name_list: 要拷贝的sheet页列表
        :return:
        '''
        for sheet_name in sheet_name_list: #[['前端功能页面-基本验证', {'D': '结算信息汇总查询-沪（Ⅰ级）'}], ['查询时段', {'D': '结算信息汇总查询-沪（Ⅰ级）', 'E': '查询时段-清算日期'}], ['单项输入框', {'D': '结算信息汇总查询-沪（Ⅰ级）', 'E': '单项输入框-挂牌年份'}], ['左右互选框', {'D': '结算信息汇总查询-沪（Ⅰ级）', 'E': '左右互选框-交收会员'}], ['左右互选框', {'D': '结算信息汇总查询-沪（Ⅰ级）', 'E': '左右互选框-结果代码'}]]
            if not self.extract_sheet(formwork_path,formwork_name,sheet_name_tag,sheet_name[0],remove_flg = False,replace_dict=sheet_name[1]):
                return False
        return True

    def extract_sheet(self, formwork_path, formwork_name, sheet_name_tag, sheet_name_src, remove_flg=True,replace_dict=None):
        '''
        sheet页拷贝，暂不支持 合并单元格的拷贝
        :param formwork_path:配置文件文档路径
        :param formwork_name:配置文件名称
        :param sheet_name_tag:要创建的sheet页名称
        :param sheet_name_src:要拷贝的sheet页名称
        :param remove_flg:重复sheet页删除标志 True-删除 False-不删除
        :param replace_dict:要替换的字段
        :return:
        '''
        if not self.read_excel_src(os.path.join(formwork_path,formwork_name)):
            return False
        ws_src = self.src_excelbook[sheet_name_src] #用例配置文件的sheet，例如 '前端功能页面-基本验证''查询时段''单项输入框''左右互选框'等等
        if sheet_name_tag not in self.tag_excelbook.sheetnames: #如果创建的sheet页名称（例如结算信息汇总查询-沪（Ⅰ级））不在用例模板文件里
            ws_tag = self.tag_excelbook.create_sheet(sheet_name_tag, -1) #实际插入的是倒数第二个位置
            self.rowbegin_tag = 0  # 写入的起始行
            rowbegin_src = 0  # 读取的起始行
        else: #创建的sheet页名称（例如结算信息汇总查询-沪（Ⅰ级））在用例模板文件里
            if remove_flg: #如果要删除用例模板excel中重复sheet页
                self.tag_excelbook.remove(self.tag_excelbook[sheet_name_tag])
                ws_tag = self.tag_excelbook.create_sheet(sheet_name_tag,-1)
                self.rowbegin_tag = 0  # 写入的起始行
                rowbegin_src = 0  # 读取的起始行
            else:
                ws_tag =  self.tag_excelbook[sheet_name_tag]
                rowbegin_src = 1  # 读取用例配置文件的起始行，如果有重复页面且不删除重复页面，则用例配置文件第一行不用拷贝，直接从第二行拷贝
        max_row = ws_src.max_row #用例配置文件的sheet，有效行数
        max_column = ws_src.max_column #用例配置文件的sheet，有效列数
        num = 0  # 源sheet页拷贝完成后的起始行偏移量
        for m_src in range(1 + rowbegin_src, max_row + 1): #循环用例配置文件sheet页的行
            # 判断每行第一列的值是否为空，如果为空则判定为尾行，不再拷贝数据与格式
            if not ws_src['A{}'.format(m_src)].value or not ws_src['A{}'.format(m_src)].value.strip():
                break
            num = m_src
            ws_tag.row_dimensions[m_src].height = ws_src.row_dimensions[m_src].height  #把用例配置sheet页中的行高 赋值给 用例模板sheet页
            for n_src in range(1, max_column + 1):#循环用例配置文件的列
                c_src = self.chr_own(n_src)
                i_src = '%s%d' % (c_src, m_src)  # 单元格编号源，例如A1 B1 C1 D1...A1 B1...A3 B3
                i_tag = '%s%d' % (c_src, m_src + self.rowbegin_tag)  # 单元格编号目标
                if m_src == 1:
                    ws_tag.column_dimensions[c_src].width = ws_src.column_dimensions[c_src].width #把用例配置sheet页中的列宽 赋值给 用例模板sheet页
                try:
                    getattr(ws_src[i_src], 'value')
                    cell_src = ws_src[i_src] # 获取用例配置文件sheet页单元格数据
                    if replace_dict and m_src > 1 and c_src in replace_dict: #replace_dict {'D': '结算信息汇总查询-沪（Ⅰ级）', 'E': '查询时段-清算日期'} #从第2行开始，拷贝DE列
                        ws_tag[i_tag].value = replace_dict[c_src]  # 拷贝并替换数据
                    else:
                        ws_tag[i_tag].value = cell_src.value  # 拷贝数据 #用例配置文件除了DE列，其他都原样拷贝 到用例模板文件
                    if cell_src.has_style:# 拷贝格式
                        ws_tag[i_tag].font = copy.copy(cell_src.font)
                        ws_tag[i_tag].border = copy.copy(cell_src.border)
                        ws_tag[i_tag].fill = copy.copy(cell_src.fill)
                        ws_tag[i_tag].number_format = copy.copy(cell_src.number_format)
                        ws_tag[i_tag].protection = copy.copy(cell_src.protection)
                        ws_tag[i_tag].alignment = copy.copy(cell_src.alignment)
                except AttributeError as e:
                    logger.warning("error:cell(%s) is %s"% (i_src,e))
                    continue

        # 设置数据有效性
        if remove_flg:
            # 复制每个sheet页的所有“数据有效性”设置
            for valid_src in ws_src.data_validations.dataValidation:
                # 复制每个sheet页的每个“数据有效性”设置
                ws_tag.add_data_validation(valid_src)
        else:
            # 复制每个sheet页的所有“数据有效性”设置
            for valid_src in ws_src.data_validations.dataValidation:
                #复制每个sheet页的每个“数据有效性”设置
                dv = DataValidation(showErrorMessage=valid_src.showErrorMessage
                                    ,showDropDown = valid_src.showDropDown
                                    ,showInputMessage = valid_src.showInputMessage
                                    ,allowBlank = valid_src.allowBlank
                                    ,errorTitle = valid_src.errorTitle
                                    ,error = valid_src.error
                                    , promptTitle = valid_src.promptTitle
                                    , prompt = valid_src.prompt
                                    ,type = valid_src.type
                                    ,errorStyle = valid_src.errorStyle
                                    ,imeMode = valid_src.imeMode
                                    ,operator = valid_src.operator
                                    ,formula1 = valid_src.formula1
                                    ,formula2 = valid_src.formula2)
                # 复制每个sheet页的每个“数据有效性”对应每个单元格的设置
                for cell_src in valid_src.sqref.ranges:
                    cell_tag_min_row = cell_src.min_row + self.rowbegin_tag
                    cell_tag_min_col = cell_src.min_col
                    cell_tag_max_row = cell_src.max_row + self.rowbegin_tag if cell_src.max_row <= num else num + self.rowbegin_tag
                    cell_tag_max_col = cell_src.max_col if cell_src.max_col <= max_column else max_column
                    # 不能add一个范围的单元格，只能一个一个的add 。。

                    for row in range(cell_tag_min_row, cell_tag_max_row + 1):
                        for col in range(cell_tag_min_col, cell_tag_max_col + 1):
                            cell_tag_col_c = self.chr_own(col)
                            dv.add(ws_tag['%s%d' % (cell_tag_col_c, row)])

                ws_tag.add_data_validation(dv)

        self.rowbegin_tag += num - 1
        self.src_excelbook.close()
        self.src_excelbook = None
        return True

    def delete_sheet(self, sheetname):
        '''
        删除目标文档指定的sheet页
        :param sheetname: 要删除的sheet页名称
        :return: True/False
        '''

        if sheetname in self.tag_excelbook.sheetnames:
            self.tag_excelbook.remove(self.tag_excelbook[sheetname])
        return True

    def model_excelbook_for_save(self, path, excelname):
        '''
        保存excel文件
        :param path:
        :param excelname:
        :return: True/False
        '''

        if self.tag_excelbook == None:
            logger.warning('请先执行[read_excel]方法')
            return False
        if path:
            if not os.path.exists(path):
                os.mkdir(path)
        excelname = os.path.join(path,excelname)

        # 保存excel文件
        try:
            self.tag_excelbook.save(excelname)
            self.tag_excelbook.close()
        except Exception as e:
            logger.error('报错信息：' + repr(e) + ',报错的行号:' + str(e.__traceback__.tb_lineno))
            return False
        return True

    def chr_own(self,num):
        '''
        sheet页的列号由数字转换为字母
        :param num:
        :return:
        '''
        if num <27:
            c_num = chr(num+64) #ASCII字符，chr(64)=‘A'

        else:
            if num < 677:
                c_num = chr(divmod(num,26)[0]+64) +chr(divmod(num,26)[1]+64)
            else:
                c_num = chr(divmod(num, 676)[0] + 64) + chr(divmod(divmod(num,676)[1],26)[0]+64) +chr(divmod(divmod(num,676)[1],26)[1]+64)
        return c_num

    def conditional_formatting(self, sheetname):
        '''
        以约定好的条件格式 进行设置
        :param sheetname:目标sheet页名称
        :return:True / False
        '''
        if not self.tag_excelbook:
            logger.warning("没有指定目标文档")
            return False

        try:
            ws_tag = self.tag_excelbook[sheetname]
            self.rule_cell1.formula[0] =self.rule_cell1.formula[0].format('K2:K%s'%(str(self.rowbegin_tag + 1)))
            self.rule_cell2.formula[0] = self.rule_cell2.formula[0].format('K2:K%s' % (str(self.rowbegin_tag + 1)))
            self.rule_cell3.formula[0] = self.rule_cell3.formula[0].format('N2:N%s' % (str(self.rowbegin_tag + 1)))
            self.rule_cell4.formula[0] = self.rule_cell4.formula[0].format('N2:N%s'%(str(self.rowbegin_tag + 1)))
            self.rule_cell5.formula[0] = self.rule_cell5.formula[0].format('N2:N%s'%(str(self.rowbegin_tag + 1)))
            self.rule_cell6.formula[0] = self.rule_cell6.formula[0].format('C2:C%s'%(str(self.rowbegin_tag + 1)))
            self.rule_cell7.formula[0] = self.rule_cell7.formula[0].format('C2:C%s'%(str(self.rowbegin_tag + 1)))
            self.rule_cell8.formula[0] = self.rule_cell8.formula[0].format('C2:C%s'%(str(self.rowbegin_tag + 1)))
            self.rule_cell9.formula[0] = self.rule_cell9.formula[0].format('B2:B%s'%(str(self.rowbegin_tag + 1)))
            ws_tag.conditional_formatting.add('K2:K%s'%(str(self.rowbegin_tag + 1)), self.rule_cell1)
            ws_tag.conditional_formatting.add('K2:K%s'%(str(self.rowbegin_tag + 1)), self.rule_cell2)
            ws_tag.conditional_formatting.add('N2:N%s'%(str(self.rowbegin_tag + 1)), self.rule_cell3)
            ws_tag.conditional_formatting.add('N2:N%s'%(str(self.rowbegin_tag + 1)), self.rule_cell4)
            ws_tag.conditional_formatting.add('N2:N%s'%(str(self.rowbegin_tag + 1)), self.rule_cell5)
            ws_tag.conditional_formatting.add('C2:C%s'%(str(self.rowbegin_tag + 1)), self.rule_cell6)
            ws_tag.conditional_formatting.add('C2:C%s'%(str(self.rowbegin_tag + 1)), self.rule_cell7)
            ws_tag.conditional_formatting.add('C2:C%s'%(str(self.rowbegin_tag + 1)), self.rule_cell8)
            ws_tag.conditional_formatting.add('B2:B%s'%(str(self.rowbegin_tag + 1)), self.rule_cell9)

            # self.rule_cell10.formula[0] = self.rule_cell10.formula[0].format('A2:A%s' % (str(self.rowbegin_tag + 1)))
            # ws_tag.conditional_formatting.add('A2:A%s'%(str(self.rowbegin_tag + 1)), self.rule_cell10)
            # self.rule_cell10.formula[0] = self.rule_cell10.formula[0].format('D2:D%s' % (str(self.rowbegin_tag + 1)))
            # ws_tag.conditional_formatting.add('D2:D%s'%(str(self.rowbegin_tag + 1)), self.rule_cell10)
            # self.rule_cell10.formula[0] = self.rule_cell10.formula[0].format('E2:E%s' % (str(self.rowbegin_tag + 1)))
            # ws_tag.conditional_formatting.add('E2:E%s'%(str(self.rowbegin_tag + 1)), self.rule_cell10)
            # self.rule_cell10.formula[0] = self.rule_cell10.formula[0].format('L2:L%s' % (str(self.rowbegin_tag + 1)))
            # ws_tag.conditional_formatting.add('L2:L%s'%(str(self.rowbegin_tag + 1)), self.rule_cell10)
            # self.rule_cell10.formula[0] = self.rule_cell10.formula[0].format('M2:M%s' % (str(self.rowbegin_tag + 1)))
            # ws_tag.conditional_formatting.add('M2:M%s'%(str(self.rowbegin_tag + 1)), self.rule_cell10)

            self.set_column_font(ws_tag,'A',self.font_cell1)
            self.set_column_font(ws_tag,'D',self.font_cell1)
            self.set_column_font(ws_tag,'E',self.font_cell1)
            self.set_column_font(ws_tag,'L',self.font_cell1)
            self.set_column_font(ws_tag,'M',self.font_cell1)


        except Exception as e:
            logger.error('报错信息：' + repr(e) + ',报错的行号:' + str(e.__traceback__.tb_lineno))
            return False
        return True

    def set_column_font(self,sheet, column_letter, font):
        """
        设置工作表中指定列的字体格式
        Args:
            sheet (Worksheet): 要设置的工作表对象
            column_letter (str): 列的字母表示，例如B列
            font (Font): 要应用的字体样式
        """
        for cell in sheet[column_letter]:
            cell.font = font