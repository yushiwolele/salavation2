# --*-- coding:UTF-8 --*--
# @Author : libeijie
# @Time : 2023/9/4 10:00

import logging
import os
import time
import threading
from django.db import connection
from collections import Counter
from openpyxl import load_workbook
from testcases.models import *
from Salvation.settings import BASE_DIR
from common.cfg_helper import *

logger = logging.getLogger('salvation')

# 创建全局锁对象
testcase_config_lock = threading.Lock() #用例配置锁
testcase_template_lock = threading.Lock() #用例模板锁

class ConfigService:
    def __init__(self):
        self.cursor = connection.cursor()
        cfg_file = ''.join([str(BASE_DIR), '/apps/data/config.cfg'])
        self.cfg_read = CfgHelper(cfg_file)


    def get_config_file(self):
        '''
        查询所有用例配置文件数据
        :return:
        '''
        testcase_config_lock.acquire()
        try:
            sql = "select id,config_file_name,config_file_type,batch_time from testcase_configfile where tags='0'"
            self.cursor.execute(sql)
            result =self.cursor.fetchall()
            config_list = self.handle_sql_data(result)
            return config_list
        finally:
            testcase_config_lock.release()

    def get_template_file(self):
        '''
        查询所有用例模板文件数据
        :return:
        '''
        testcase_template_lock.acquire()
        try:
            sql = "select id,config_file_name,config_file_type,batch_time from testcase_configfile where tags='1'"
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            config_list = self.handle_sql_data(result)
            return config_list
        finally:
            testcase_template_lock.release()

    def handle_sql_data(self,result):
        '''
        处理数据库数据为固定格式
        :param result: 数据库数据集合
        :return:
        '''
        config_list = list()
        for item in result:
            config_id = item[0]
            config_name = item[1]
            config_type = item[2]
            if item[3]:
                batch_time = item[3].strftime("%Y-%m-%d %H:%M:%S")
            else:
                batch_time = None
            config_list.append({'config_id': config_id, 'config_name': config_name, 'config_type': config_type,
                                'batch_time': batch_time})
        return config_list

    def put_file_insert_into_db(self,file,file_type,insert_update_flag,testcase_config,tags):
        '''
        将上传的文件放置在服务器上，并且将文件名及路径保存到数据库
        :param file: 上传的文件
        :param file_type: 上传的文件类型
        :param insert_update_flag: 0-插入；1-更新
        :param testcase_config: 已存在此文件类型的实例
        :return:
        '''
        #上传至服务器
        put_result = self.put_file(file,tags)
        put_result_code = put_result['code']
        if put_result_code == 0:  # 成功放在了服务器的指定路径上
            file_path = put_result['message']
            create_time = time.strftime('%Y-%m-%dT%H:%M:%S', time.localtime(time.time()))
            file_name = file.name
            if insert_update_flag == 1: #更新数据库
                testcase_config.update(config_file_name=file_name,config_file_path=file_path,batch_time=create_time)
            else: #插入数据库
                TestCaseConfigModel.objects.create(config_file_name=file_name,config_file_path=file_path,config_file_type=file_type,batch_time=create_time,tags=tags)
            logger.info({'config_file_name':file_name,'config_file_path':file_path})
        return put_result

    def put_file(self,file,tags):
        '''
        将文件放置在服务器上
        :param file: 上传的文件
        :param tags: 文件标志；0-配置文件；1-模板文件
        :return: dict
        '''
        if tags == '0':
            file_path = self.cfg_read.read_cfg_get_value_by_key('config_file', 'file_config_path')
        else:
            file_path = self.cfg_read.read_cfg_get_value_by_key('config_file', 'file_template_path')
        try:
            # 如果路径不存在，则创建
            if not os.path.exists(file_path):
                os.makedirs(file_path)
            #构建完整路径
            destination_file_path = os.path.join(file_path,file.name)
            #保存文件到目标路径，以写入二进制模式打开文件
            with open(destination_file_path,'wb') as destination_file:
                #shutil.copyfileobj(file,destination_file) #使用此方法一直无法正确拷贝文件，总是提示文件损坏
                for chunk in file.chunks():
                    destination_file.write(chunk)
            return {'code': 0,'message':str(file.name)}
        except Exception as e:
            logger.error('报错信息：'+repr(e) + ',报错的行号:' + str(e.__traceback__.tb_lineno))
            return {'code': 1, 'message': '文件上传失败，请联系管理员！'}

    def judge_config_file_content(self,file,file_type):
        '''
        判断文件内容是否符合固定格式 (django中没有装饰器可以实现此功能)
        :param upload_file:文件
        :param file_type:文件类型
        :return:
        '''
        code = 0
        message = None
        #读取文件sheet页名字
        workbook = load_workbook(file,read_only=True)
        sheet_names_list = workbook.sheetnames
        #获取cfg定义的sheet页名
        cfg_sheet_names = self.cfg_read.read_cfg_get_value_by_key(file_type, 'sheet_names')
        cfg_sheet_names = cfg_sheet_names.split(';')
        #获取cfg定义的列首名
        col_names = self.cfg_read.read_cfg_get_value_by_key('public_name', 'beginning_name')
        col_names = (col_names.replace("\\n","\n")).split(';')
        sheet_names = [x for x in sheet_names_list if (x != '文档修改记录' and x!='目录')]
        if Counter(sheet_names) == Counter(cfg_sheet_names):
            #比较每个sheet页中的列首名是否符合cfg定义的
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                #first_row = list(list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0])
                #first_row = list(list(sheet.iter_rows(min_row=1, max_row=1))[0])

                # 获取第一行的数据的非空单元格数量
                first_row_data_count = 0
                for cell in sheet[1]:
                    if cell.value is not None:
                        first_row_data_count += 1

                if first_row_data_count == len(col_names):
                    continue
                else:
                    code = 1
                    message = '{0}格式不符合规范，请检查！'.format(sheet_name)
                    break
        else: #上传的sheet页和固定格式不符合
            code = 1
            message = '控件名称不符合规范，请检查！'
        workbook.close()
        return {'code': code, 'message': message}

    def judge_template_file_content(self,file):
        '''
        模板文件是否符合固定格式（仅判断sheet页，不判断列首名）
        :param file: 上传的文件
        :return:
        '''
        code = 0
        message = None
        #读取sheet页
        workbook = load_workbook(file, read_only=True)
        sheet_names_list = workbook.sheetnames #包含隐藏的sheet页
        # 获取cfg定义的sheet页名
        cfg_sheet_names = self.cfg_read.read_cfg_get_value_by_key('TEMPLATE', 'sheet_names')
        cfg_sheet_names = cfg_sheet_names.split(';')
        #判断sheet_names_list 是否完全包含cfg_sheet_names
        is_contained = set(cfg_sheet_names).issubset(set(sheet_names_list))
        if not is_contained:
            code = 1
            message = '模板文件格式不符合规范，请检查！'
        workbook.close()
        return {'code': code, 'message': message}

    def upload_config_file(self,file,file_type,insert_update_flag,tags):
        '''
        上传配置文件。检查文件是否存在数据库，是否符合固定格式，并放置在服务器及插入数据库
        :param file: 上传的文件
        :param file_type: 上传的文件类型
        :param insert_update_flag: 对文件的操作0-插入；1-更新
        :param tags: 文件标识，0-配置文件；1-模板文件
        :return:
        '''
        #检查文件是否符合规范
        judge_result = self.judge_config_file_content(file, file_type)

        if judge_result['code'] == 0: #符合规范
            file_path_root = self.cfg_read.read_cfg_get_value_by_key('config_file', 'file_config_path')
            testcase_config_lock.acquire()  # 加锁
            try:
                # 检查文件类型否在数据库存在
                testcase_config = self.check_config_file_by_type(file_type)
                if testcase_config:  # 存在，则update
                    # 获取服务器文件路径，并删除旧文件
                    old_file_path = os.path.join(file_path_root,testcase_config[0].config_file_path)
                    if os.path.exists(old_file_path):
                        os.remove(old_file_path)
                    insert_update_flag = 1
                put_result = self.put_file_insert_into_db(file, file_type,insert_update_flag,testcase_config,tags)
            finally:
                testcase_config_lock.release()
            return put_result

        else:
            return judge_result

    def upload_template_file(self, file, file_type, insert_update_flag, tags,template_name):
        '''
        检查模板文件是否存在数据库，是否符合固定格式，并放置在服务器及插入数据库
        :param file:上传的文件
        :param file_type:上传的文件类型
        :param insert_update_flag:对文件的操作0-插入；1-更新
        :param tags:文件标识，0-配置文件；1-模板文件
        :param template_name:旧文件名字，只有更新操作才有值
        :return:
        '''
        judge_result = self.judge_template_file_content(file)

        if judge_result['code'] == 0:  # 符合规范
            file_path_root = self.cfg_read.read_cfg_get_value_by_key('config_file', 'file_template_path')
            testcase_template_lock.acquire()
            try:
                testcase_config = self.check_template_file_by_name_tags(template_name, tags)
                if testcase_config:#更新操作
                    if template_name == file.name: #本数据更新操作
                        old_file_path = os.path.join(file_path_root,testcase_config[0].config_file_path)
                        if os.path.exists(old_file_path):
                            os.remove(old_file_path)
                        put_result = self.put_file_insert_into_db(file, file_type, insert_update_flag, testcase_config,tags)
                    else:
                        #查看上传的file_name是否在数据库中存在
                        upload_file = self.check_template_file_by_name_tags(file.name, tags)
                        if upload_file:#存在，报错
                            put_result = {'code': 1, 'message': '文件上传失败，上传的文件名字已在列表中，请选择对应文件进行更新！'}
                        else:
                            old_file_path = os.path.join(file_path_root,testcase_config[0].config_file_path)
                            if os.path.exists(old_file_path):
                                os.remove(old_file_path)
                            put_result = self.put_file_insert_into_db(file, file_type, insert_update_flag, testcase_config,tags)
                else: #插入操作
                    upload_file = self.check_template_file_by_name_tags(file.name, tags)
                    if upload_file:#要插入的文件已在数据库存在，则更新
                        upload_file_path = os.path.join(file_path_root,upload_file[0].config_file_path)
                        if os.path.exists(upload_file_path):
                            os.remove(upload_file_path)
                        insert_update_flag = 1
                    put_result = self.put_file_insert_into_db(file, file_type, insert_update_flag, upload_file,tags)
            finally:
                testcase_template_lock.release()
        else:
            put_result = judge_result
        return put_result


    def check_template_file_by_name_tags(self,file_name,tags):
        '''
        根据文件name和文件标识获取数据
        :param file: 上传的文件
        :param tags: 文件标识
        :return:
        '''
        testcase_config = TestCaseConfigModel.objects.filter(config_file_name=file_name,tags=tags)
        count = testcase_config.count()
        if count >= 1:  # 数据库有数据
            return testcase_config
        else:  # 无数据
            return None


    def check_config_file_by_type(self,file_type):
        '''
        检查模板文件类型是否在数据库中存在
        :param file_type: 上传的文件类型
        :return: 存在 返回数据库数据/不存在 返回None
        '''
        testcase_config = TestCaseConfigModel.objects.filter(config_file_type=file_type)
        count = testcase_config.count()
        if count >= 1: #数据库有数据
            return testcase_config
        else: #无数据
            return None

    def download_config_file(self,config_flag,tags):
        '''
        下载用例配置文件
        :param config_flag: 文件类型
        :param tags:文件标识，0-配置文件，1-模板文件
        :return:
        '''
        file_path_root = self.cfg_read.read_cfg_get_value_by_key('config_file', 'file_config_path')
        testcase_config_lock.acquire()  # 加锁
        try:
            file_model = TestCaseConfigModel.objects.filter(config_file_type=config_flag,tags=tags)
            if file_model:
                file_path = os.path.join(file_path_root,file_model[0].config_file_path)
                file_name = file_model[0].config_file_name
            else:
                file_path = None
                file_name = None
            return file_path, file_name
        finally:
            testcase_config_lock.release()

    def download_template_file(self, config_name, tags):
        '''
        下载用例模板文件
        :param config_name: 文件名字
        :param tags:文件标识，0-配置文件，1-模板文件
        :return:
        '''
        file_path_root = self.cfg_read.read_cfg_get_value_by_key('config_file', 'file_template_path')
        testcase_template_lock.acquire()  # 加锁
        try:
            file_model = TestCaseConfigModel.objects.filter(config_file_name=config_name, tags=tags)
            if file_model:
                file_path = os.path.join(file_path_root, file_model[0].config_file_path)
                file_name = file_model[0].config_file_name
            else:
                file_path = None
                file_name = None
            return file_path, file_name
        finally:
            testcase_template_lock.release()

    def delete_template_testcase(self,template_name):
        '''
        删除文件与数据
        :return:
        '''
        testcase_template_lock.acquire()
        try:
            queryset = TestCaseConfigModel.objects.filter(config_file_name=template_name)
            #删除物理文件
            file_path_root = self.cfg_read.read_cfg_get_value_by_key('config_file', 'file_template_path')
            config_path = os.path.join(file_path_root,queryset[0].config_file_path)
            if os.path.exists(config_path):
                os.remove(config_path)
            #删除数据库数据
            queryset.delete()
            return {'code':0,'message':'文件删除成功！'}
        except Exception as e:
            logger.error('报错信息：' + repr(e) + ',报错的行号:' + str(e.__traceback__.tb_lineno))
            return {'code': 1, 'message': '文件删除失败，请联系管理员！'}
        finally:
            testcase_template_lock.release()

    def delete_config_testcase(self,config_type):
        '''
        删除用例配置文件
        :param config_type: 用例配置文件类型
        :return:
        '''
        testcase_config_lock.acquire()
        try:
            queryset = TestCaseConfigModel.objects.filter(config_file_type=config_type)
            # 删除物理文件
            file_path_root = self.cfg_read.read_cfg_get_value_by_key('config_file', 'file_config_path')
            config_path = os.path.join(file_path_root,queryset[0].config_file_path)
            if os.path.exists(config_path):
                os.remove(config_path)
            # 删除数据库数据
            queryset.delete()
            return {'code': 0, 'message': '文件删除成功！'}
        except Exception as e:
            logger.error('报错信息：' + repr(e) + ',报错的行号:' + str(e.__traceback__.tb_lineno))
            return {'code': 1, 'message': '文件删除失败，请联系管理员！'}
        finally:
            testcase_config_lock.release()
